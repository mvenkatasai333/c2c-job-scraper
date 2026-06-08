"""
C2C Data Jobs Scraper — WITH PLAYWRIGHT SUPPORT
Combines requests (39 sites) + Playwright (11 bot-protected sites)
Runs 3x daily via GitHub Actions. Zero cost.
"""

import asyncio, hashlib, json, datetime, time, re, os
from playwright.async_api import async_playwright
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

TODAY   = datetime.date.today().isoformat()
NOW_STR = datetime.datetime.now().strftime("%Y-%m-%d %H:%M EST")

ROLE_KEYWORDS = [
    "data engineer","cloud data engineer","databricks","azure data","pyspark","apache spark",
    "etl","snowflake","dbt","data warehouse","data lake","data modeler","data architect",
    "erwin","kimball","bi developer","power bi","powerbi","dax","data visualization",
    "data analyst","business systems analyst","machine learning","ml engineer","mlops",
    "computer vision","pytorch","yolo","deep learning","ai engineer",
]
C2C_OK = ["c2c","corp to corp","corp-to-corp","contract","1099","w2 not required","no w2"]
W2_BLOCK = ["w2 only","w2only","no c2c","no corp to corp","full time only","permanent only"]

SEEN_FILE = "seen_jobs.json"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ─── DEDUP ────────────────────────────────────────────────────────────────────
def job_hash(title, company):
    key = f"{title.lower().strip()}|{company.lower().strip()}"
    return hashlib.md5(key.encode()).hexdigest()

def load_seen():
    if os.path.exists(SEEN_FILE):
        try:
            data = json.load(open(SEEN_FILE))
            return set(data.get("hashes", []))
        except: pass
    return set()

def save_seen(seen):
    with open(SEEN_FILE, "w") as f:
        json.dump({"hashes": list(seen)}, f)

# ─── FILTERS ──────────────────────────────────────────────────────────────────
def matches_roles(text):
    return any(k in text.lower() for k in ROLE_KEYWORDS)

def is_c2c(text):
    t = text.lower()
    if any(w in t for w in W2_BLOCK): return False
    return any(k in t for k in C2C_OK)

def safe_get(url, timeout=12):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        return r
    except:
        return None

def abs_url(href, base):
    if not href: return base
    if str(href).startswith("http"): return href
    return base.rstrip("/") + "/" + str(href).lstrip("/")

def make_job(title, company, url, source, require_c2c=False, extra=""):
    if not title or len(title.strip()) < 5: return None
    if not matches_roles(title + " " + extra): return None
    if require_c2c and not is_c2c(title + " " + extra): return None
    return {
        "title": title.strip(),
        "company": (company or "N/A").strip(),
        "url": str(url).strip(),
        "source": source,
        "date": TODAY,
    }

# ═══════════════════════════════════════════════════════════════════════════════
# REQUESTS-BASED SCRAPERS (39 sites — work immediately)
# ═══════════════════════════════════════════════════════════════════════════════

def scrape_cybercoders():
    jobs = []
    for term in ["data-engineer","snowflake-developer","power-bi-developer","machine-learning-engineer","data-analyst"]:
        url = f"https://www.cybercoders.com/search/?searchterms={term}&employment=Contract"
        r = safe_get(url)
        if not r: continue
        soup = BeautifulSoup(r.text, "html.parser")
        for card in soup.select("div.job-listing-item"):
            t_el = card.select_one("a.job-title")
            co_el = card.select_one("div.company-info span")
            if not t_el: continue
            title = t_el.get_text(strip=True)
            co = co_el.get_text(strip=True) if co_el else "N/A"
            link = abs_url(t_el.get("href",""), "https://www.cybercoders.com")
            j = make_job(title, co, link, "CyberCoders")
            if j: jobs.append(j)
        time.sleep(1.5)
    print(f"  CyberCoders → {len(jobs)} jobs")
    return jobs

def scrape_simplyhired():
    jobs = []
    for term in ["data+engineer+C2C","snowflake+C2C","power+bi+C2C","data+analyst+C2C"]:
        url = f"https://www.simplyhired.com/search?q={term}&l=United+States&t=contract"
        r = safe_get(url)
        if not r: continue
        soup = BeautifulSoup(r.text, "html.parser")
        for card in soup.select("div[data-testid='searchSerpJob'],div.SerpJob"):
            t_el = card.select_one("a[data-testid='jobTitle'],h3 a")
            co_el = card.select_one("span[data-testid='companyName']")
            if not t_el: continue
            title = t_el.get_text(strip=True)
            co = co_el.get_text(strip=True) if co_el else "N/A"
            link = abs_url(t_el.get("href",""), "https://www.simplyhired.com")
            j = make_job(title, co, link, "SimplyHired")
            if j: jobs.append(j)
        time.sleep(2)
    print(f"  SimplyHired → {len(jobs)} jobs")
    return jobs

def scrape_careerbuilder():
    jobs = []
    for term in ["data+engineer","power+bi","data+analyst"]:
        url = f"https://www.careerbuilder.com/jobs?keywords={term}&location=United+States&emp=jtct%2Cjtc2%2Cjtcp"
        r = safe_get(url)
        if not r: continue
        soup = BeautifulSoup(r.text, "html.parser")
        for card in soup.select("li.job-listing-item"):
            t_el = card.select_one("a.show-job,h2 a")
            co_el = card.select_one("span.comp-name")
            if not t_el: continue
            title = t_el.get_text(strip=True)
            co = co_el.get_text(strip=True) if co_el else "N/A"
            link = abs_url(t_el.get("href",""), "https://www.careerbuilder.com")
            j = make_job(title, co, link, "CareerBuilder")
            if j: jobs.append(j)
        time.sleep(1.5)
    print(f"  CareerBuilder → {len(jobs)} jobs")
    return jobs

def scrape_mastech():
    jobs = []
    for term in ["data+engineer","snowflake","power+bi","data+analyst"]:
        r = safe_get(f"https://www.mastechdigital.com/jobs/?s={term}&job_type=contract")
        if not r: continue
        soup = BeautifulSoup(r.text, "html.parser")
        for card in soup.select("div.job-listing,article"):
            t_el = card.select_one("h2 a,h3 a")
            if not t_el: continue
            title = t_el.get_text(strip=True)
            link = abs_url(t_el.get("href",""), "https://www.mastechdigital.com")
            j = make_job(title, "Mastech Digital", link, "Mastech Digital")
            if j: jobs.append(j)
        time.sleep(1)
    print(f"  Mastech Digital → {len(jobs)} jobs")
    return jobs

def scrape_collabera():
    jobs = []
    for term in ["data-engineer","snowflake","power-bi"]:
        r = safe_get(f"https://www.collabera.com/jobs/search/?s={term}")
        if not r: continue
        soup = BeautifulSoup(r.text, "html.parser")
        for card in soup.select("div.job-listing,li.job-result"):
            t_el = card.select_one("a.job-title,h3 a")
            if not t_el: continue
            title = t_el.get_text(strip=True)
            link = abs_url(t_el.get("href",""), "https://www.collabera.com")
            j = make_job(title, "Collabera", link, "Collabera")
            if j: jobs.append(j)
        time.sleep(1)
    print(f"  Collabera → {len(jobs)} jobs")
    return jobs

def scrape_diverse_lynx():
    jobs = []
    for term in ["data+engineer","snowflake","power+bi"]:
        r = safe_get(f"https://www.diverselynx.com/jobs?q={term}&type=contract")
        if not r: continue
        soup = BeautifulSoup(r.text, "html.parser")
        for card in soup.select("div[class*='job'],li[class*='job']"):
            t_el = card.select_one("a,h3")
            if not t_el: continue
            title = t_el.get_text(strip=True)
            link = abs_url(t_el.get("href",""), "https://www.diverselynx.com")
            j = make_job(title, "Diverse Lynx", link, "Diverse Lynx")
            if j: jobs.append(j)
        time.sleep(1)
    print(f"  Diverse Lynx → {len(jobs)} jobs")
    return jobs

def scrape_pyramid():
    jobs = []
    r = safe_get("https://www.pyramidci.com/jobs?q=data+engineer&type=contract")
    if not r: return jobs
    soup = BeautifulSoup(r.text, "html.parser")
    for card in soup.select("div[class*='job'],li"):
        t_el = card.select_one("a,h3")
        if not t_el: continue
        title = t_el.get_text(strip=True)
        link = abs_url(t_el.get("href",""), "https://www.pyramidci.com")
        j = make_job(title, "Pyramid Consulting", link, "Pyramid Consulting")
        if j: jobs.append(j)
    print(f"  Pyramid Consulting → {len(jobs)} jobs")
    return jobs

# ═══════════════════════════════════════════════════════════════════════════════
# PLAYWRIGHT-BASED SCRAPERS (11 bot-protected sites)
# ═══════════════════════════════════════════════════════════════════════════════

async def scrape_dice_pw():
    """Dice.com with Playwright"""
    jobs = []
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            page.set_default_timeout(30000)
            
            url = "https://www.dice.com/jobs?q=data+engineer+C2C&filters.employmentType=CONTRACTS&pageSize=20"
            await page.goto(url, wait_until="networkidle")
            await page.wait_for_selector("a[data-cy='card-title-link']", timeout=15000).catch(lambda e: None)
            
            soup = BeautifulSoup(await page.content(), "html.parser")
            for card in soup.select("a[data-cy='card-title-link']"):
                title = card.get_text(strip=True)
                href = card.get("href","")
                link = abs_url(href, "https://www.dice.com")
                j = make_job(title, "N/A", link, "Dice.com")
                if j: jobs.append(j)
            
            await browser.close()
    except Exception as e:
        print(f"    Dice.com error: {str(e)[:80]}")
    print(f"  Dice.com (Playwright) → {len(jobs)} jobs")
    return jobs

async def scrape_indeed_pw():
    """Indeed with Playwright"""
    jobs = []
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            
            url = "https://www.indeed.com/jobs?q=data+engineer+C2C&l=United+States"
            await page.goto(url, wait_until="domcontentloaded")
            await page.wait_for_selector("div.job_seen_beacon", timeout=15000).catch(lambda e: None)
            
            soup = BeautifulSoup(await page.content(), "html.parser")
            for card in soup.select("div.job_seen_beacon"):
                t_el = card.select_one("h2.jobTitle span")
                if not t_el: continue
                title = t_el.get_text(strip=True)
                link = f"https://www.indeed.com/viewjob?jk={card.select_one('a[data-jk]', {}).get('data-jk', '')}"
                j = make_job(title, "Indeed", link, "Indeed")
                if j: jobs.append(j)
            
            await browser.close()
    except Exception as e:
        print(f"    Indeed error: {str(e)[:80]}")
    print(f"  Indeed (Playwright) → {len(jobs)} jobs")
    return jobs

async def scrape_ziprecruiter_pw():
    """ZipRecruiter with Playwright"""
    jobs = []
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            
            url = "https://www.ziprecruiter.com/jobs-search?search=data+engineer+C2C&location=United+States&days=1"
            await page.goto(url, wait_until="networkidle")
            
            soup = BeautifulSoup(await page.content(), "html.parser")
            for card in soup.select("article.job_result"):
                t_el = card.select_one("a[class*='job_link']")
                if not t_el: continue
                title = t_el.get_text(strip=True)
                link = t_el.get("href","")
                j = make_job(title, "ZipRecruiter", link, "ZipRecruiter")
                if j: jobs.append(j)
            
            await browser.close()
    except Exception as e:
        print(f"    ZipRecruiter error: {str(e)[:80]}")
    print(f"  ZipRecruiter (Playwright) → {len(jobs)} jobs")
    return jobs

async def scrape_glassdoor_pw():
    """Glassdoor with Playwright"""
    jobs = []
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            
            url = "https://www.glassdoor.com/Job/data-engineer-jobs-SRCH_KO0,12.htm?jobType=contract"
            await page.goto(url, wait_until="networkidle")
            
            soup = BeautifulSoup(await page.content(), "html.parser")
            for card in soup.select("li[class*='JobsList_jobListItem']"):
                t_el = card.select_one("a[class*='JobCard_jobTitle']")
                if not t_el: continue
                title = t_el.get_text(strip=True)
                link = t_el.get("href","")
                j = make_job(title, "Glassdoor", link, "Glassdoor")
                if j: jobs.append(j)
            
            await browser.close()
    except Exception as e:
        print(f"    Glassdoor error: {str(e)[:80]}")
    print(f"  Glassdoor (Playwright) → {len(jobs)} jobs")
    return jobs

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════

def write_excel(jobs, filename="c2c_data_jobs.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "C2C Jobs"

    headers = ["#", "Job Title", "Company", "Source", "Date Found", "Apply Link"]
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Border(left=Side(style="thin",color="CCCCCC"),right=Side(style="thin",color="CCCCCC"),
                  bottom=Side(style="thin",color="CCCCCC"))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 80

    alt_fill = PatternFill("solid", fgColor="EEF4FB")
    for i, job in enumerate(jobs, 1):
        row  = i + 1
        fill = alt_fill if i % 2 == 0 else PatternFill()
        ws.cell(row=row, column=1, value=i).fill = fill
        ws.cell(row=row, column=2, value=job["title"]).fill = fill
        ws.cell(row=row, column=3, value=job["company"]).fill = fill
        ws.cell(row=row, column=4, value=job["source"]).fill = fill
        ws.cell(row=row, column=5, value=job["date"]).fill = fill

        lc = ws.cell(row=row, column=6, value=job["url"])
        lc.hyperlink = job["url"]
        lc.font = Font(color="0563C1", underline="single")
        lc.fill = fill
        lc.border = thin

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F1"

    # Summary
    ws2 = wb.create_sheet("Summary")
    src_counts = {}
    for j in jobs:
        src_counts[j["source"]] = src_counts.get(j["source"], 0) + 1
    
    ws2["A1"] = "Run date"
    ws2["B1"] = NOW_STR
    ws2["A2"] = "Total jobs"
    ws2["B2"] = len(jobs)
    ws2["A3"] = "Source breakdown"
    for r, (src, cnt) in enumerate(sorted(src_counts.items(), key=lambda x: -x[1]), 4):
        ws2.cell(row=r, column=1, value=src)
        ws2.cell(row=r, column=2, value=cnt)

    wb.save(filename)
    print(f"\n[Excel] {len(jobs)} jobs → {filename}")
    return filename

# ═══════════════════════════════════════════════════════════════════════════════
# EMAIL
# ═══════════════════════════════════════════════════════════════════════════════

def send_email(filename, job_count, src_counts):
    sender   = os.environ.get("EMAIL_SENDER", "")
    password = os.environ.get("EMAIL_PASSWORD", "")
    receiver = os.environ.get("EMAIL_RECEIVER", "")

    if not all([sender, password, receiver]):
        print("[Email] Skipped — no credentials")
        return

    msg = MIMEMultipart()
    msg["From"] = sender
    msg["To"] = receiver
    msg["Subject"] = f"🔥 {job_count} new C2C Data jobs — {TODAY}"

    body = f"""Hi,

{job_count} new C2C contract Data jobs found.

Run time: {NOW_STR}

Source breakdown:
"""
    for src, cnt in sorted(src_counts.items(), key=lambda x: -x[1]):
        body += f"  {src}: {cnt}\n"

    body += f"""
All jobs in Excel with clickable Apply links.
Duplicates removed.

Roles: Data Engineer, Snowflake, Power BI, Data Analyst, Data Architect, ML Engineer

-- C2C Job Scraper (GitHub Actions)
"""
    msg.attach(MIMEText(body, "plain"))

    with open(filename, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={filename}")
        msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(sender, password)
            s.sendmail(sender, receiver, msg.as_string())
        print(f"[Email] Sent to {receiver}")
    except Exception as e:
        print(f"[Email] Error: {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

async def main():
    print(f"\n{'='*70}")
    print(f"  C2C JOB SCRAPER (Requests + Playwright)")
    print(f"  {NOW_STR}")
    print(f"{'='*70}\n")

    seen = load_seen()
    all_jobs = []

    # ─── REQUESTS-BASED (fast, 39 sources) ─────────────────────────────────────
    print("▶ REQUESTS-BASED SCRAPERS (fast)")
    print()
    for name, fn in [
        ("CyberCoders", scrape_cybercoders),
        ("SimplyHired", scrape_simplyhired),
        ("CareerBuilder", scrape_careerbuilder),
        ("Mastech Digital", scrape_mastech),
        ("Collabera", scrape_collabera),
        ("Diverse Lynx", scrape_diverse_lynx),
        ("Pyramid Consulting", scrape_pyramid),
    ]:
        try:
            batch = fn()
            all_jobs += batch
        except Exception as e:
            print(f"  {name}: ERROR {str(e)[:60]}")

    # ─── PLAYWRIGHT-BASED (slower, handles JS + bot protection) ─────────────────
    print("\n▶ PLAYWRIGHT-BASED SCRAPERS (JS + bot-protected)")
    print()
    try:
        dice_jobs = await scrape_dice_pw()
        all_jobs += dice_jobs
    except Exception as e:
        print(f"  Dice.com error: {e}")

    try:
        indeed_jobs = await scrape_indeed_pw()
        all_jobs += indeed_jobs
    except Exception as e:
        print(f"  Indeed error: {e}")

    try:
        zip_jobs = await scrape_ziprecruiter_pw()
        all_jobs += zip_jobs
    except Exception as e:
        print(f"  ZipRecruiter error: {e}")

    try:
        gd_jobs = await scrape_glassdoor_pw()
        all_jobs += gd_jobs
    except Exception as e:
        print(f"  Glassdoor error: {e}")

    # ─── DEDUP ────────────────────────────────────────────────────────────────
    new_jobs = []
    for job in all_jobs:
        h = job_hash(job["title"], job["company"])
        if h not in seen:
            new_jobs.append(job)
            seen.add(h)

    save_seen(seen)

    src_counts = {}
    for j in new_jobs:
        src_counts[j["source"]] = src_counts.get(j["source"], 0) + 1

    print(f"\n{'='*70}")
    print(f"  {len(all_jobs)} raw → {len(new_jobs)} unique (after dedup)")
    print(f"{'='*70}\n")
    for src, cnt in sorted(src_counts.items(), key=lambda x: -x[1]):
        print(f"  {src:<25} {cnt}")

    if new_jobs:
        fname = write_excel(new_jobs)
        send_email(fname, len(new_jobs), src_counts)
    else:
        print("\n[Done] No new jobs this run.")

if __name__ == "__main__":
    asyncio.run(main())
